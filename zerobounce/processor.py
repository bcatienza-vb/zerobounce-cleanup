import os
import sys
import pandas as pd
import numbers
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

from .config import load_config


def read_csv_safe(file_path):
    try:
        return pd.read_csv(file_path, encoding="utf-8", dtype=str, low_memory=False)
    except UnicodeDecodeError:
        return pd.read_csv(file_path, encoding="latin1", dtype=str, low_memory=False)


def adjust_format_excel(excel_path, width=20, number_format_threshold=7):
    if not os.path.exists(excel_path):
        return

    try:
        workbook = load_workbook(excel_path)
        no_border = Border(left=None, right=None, top=None, bottom=None)
        integer_format = "0"

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            if worksheet.max_row == 0 or worksheet.max_column == 0:
                continue

            if worksheet.max_row >= 1:
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.border = no_border

            for col_idx in range(1, worksheet.max_column + 1):
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].width = width

                if worksheet.max_row >= 2:
                    for row_idx in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        current_value = cell.value

                        convert_to_number = False
                        numeric_value = None

                        if isinstance(current_value, str) and current_value.isdigit():
                            if len(current_value) > number_format_threshold:
                                try:
                                    numeric_value = int(current_value)
                                    convert_to_number = True
                                except (ValueError, TypeError):
                                    pass
                        elif isinstance(current_value, numbers.Number) and not isinstance(current_value, bool):
                            try:
                                if isinstance(current_value, float):
                                    num_int_part = int(current_value)
                                else:
                                    num_int_part = current_value

                                if len(str(abs(num_int_part))) > number_format_threshold:
                                    numeric_value = current_value
                                    convert_to_number = True
                            except (ValueError, TypeError):
                                pass

                        if convert_to_number and numeric_value is not None:
                            cell.value = numeric_value
                            cell.number_format = integer_format

        workbook.save(excel_path)
    except Exception as e:
        print(f"Warning: Could not apply formatting to {os.path.basename(excel_path)}. Error: {e}")


def clean_csv_to_xlsx(input_csv, output_xlsx=None, log=print):
    config = load_config()
    keep_cols = config["columns"]
    column_order = config["order"]

    df = read_csv_safe(input_csv)
    df.columns = df.columns.astype(str)

    df = df[[col for col in df.columns if col in keep_cols]]
    df = df.rename(columns=keep_cols)
    df = df[[col for col in column_order if col in df.columns]]

    if "Phone Number" in df.columns:
        df["Phone Number"] = (
            df["Phone Number"]
            .fillna("")
            .astype(str)
            .str.replace("\t", "", regex=False)
            .str.replace(r"(?!^)\+", "", regex=True)
            .str.replace(r"[^\d+]", "", regex=True)
        )
        df["Phone Number"] = df["Phone Number"].apply(lambda x: f"'{x}" if x else x)

    if not output_xlsx:
        base, _ = os.path.splitext(input_csv)
        output_xlsx = base + ".xlsx"

    df.to_excel(output_xlsx, index=False)
    adjust_format_excel(output_xlsx, width=20)

    log(f"Cleaned file saved as: {output_xlsx}")
    return output_xlsx


def process_zerobounce(file_path, log=print):
    if not os.path.isfile(file_path):
        log(f"Error: File not found at {file_path}")
        return None

    if not file_path.lower().endswith(".csv"):
        log(f"Error: File is not a CSV: {file_path}")
        return None

    log(f"Processing: {os.path.basename(file_path)}")
    df = read_csv_safe(file_path)

    required_cols = ["ZB Status", "ZB Sub status"]
    if not all(col in df.columns for col in required_cols):
        missing = [col for col in required_cols if col not in df.columns]
        log(f"Error: Missing columns: {', '.join(missing)}")
        return None

    df["ZB Status"] = df["ZB Status"].fillna("").astype(str).str.lower().str.strip()
    df["ZB Sub status"] = df["ZB Sub status"].fillna("").astype(str).str.lower().str.strip()

    status_col = df["ZB Status"]
    sub_status_col = df["ZB Sub status"]

    cond1 = (status_col == "do_not_mail") & (sub_status_col.isin(["role_based_catch_all", "role_based"]))
    cond2 = status_col == "valid"
    cond3 = (status_col == "catch-all") | (sub_status_col == "catch_all")

    is_catch_all = cond1 | cond3
    is_valid = cond2 & ~is_catch_all
    is_invalid = ~(is_valid | is_catch_all)

    valid_df = df[is_valid].copy()
    catch_all_df = df[is_catch_all].copy()
    invalid_df = df[is_invalid].copy()

    log(f"  Valid: {len(valid_df)}, Catch-All: {len(catch_all_df)}, Invalid: {len(invalid_df)}")

    cols_to_remove = ["ZB Status", "ZB Sub status"]
    valid_df.drop(columns=cols_to_remove, errors="ignore", inplace=True)
    catch_all_df.drop(columns=cols_to_remove, errors="ignore", inplace=True)
    invalid_df.drop(columns=cols_to_remove, errors="ignore", inplace=True)

    stats_df = pd.DataFrame({
        "Valid": [len(valid_df)],
        "Catch-All": [len(catch_all_df)],
        "Invalid": [len(invalid_df)],
        "Total": [len(df)],
    })

    input_dir = os.path.dirname(file_path)
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    for pattern in ["all_results", "ALL_RESULTS"]:
        base_name = base_name.replace(pattern, "")
    base_name = base_name.strip(" _-") or "Processed_Data"

    cleaned_dir = os.path.join(input_dir, "cleaned")
    os.makedirs(cleaned_dir, exist_ok=True)

    paths = {
        "combined": os.path.join(cleaned_dir, f"{base_name}_all_results.xlsx"),
        "valid": os.path.join(cleaned_dir, f"{base_name}_Valid.xlsx"),
        "catch_all": os.path.join(cleaned_dir, f"{base_name}_Catch-All.xlsx"),
        "valid_catch_all": os.path.join(cleaned_dir, f"{base_name}_Valid_and_Catch-All.xlsx"),
    }

    try:
        with pd.ExcelWriter(paths["combined"], engine="openpyxl") as writer:
            valid_df.to_excel(writer, sheet_name="Valid", index=False)
            catch_all_df.to_excel(writer, sheet_name="Catch-All", index=False)
            invalid_df.to_excel(writer, sheet_name="Invalid", index=False)
            stats_df.to_excel(writer, sheet_name="Statistics", index=False)
        adjust_format_excel(paths["combined"])
        log(f"  Saved: {os.path.basename(paths['combined'])}")
    except Exception as e:
        log(f"  Error saving combined file: {e}")

    for key, sheet, data in [
        ("valid", "Valid", valid_df),
        ("catch_all", "Catch-All", catch_all_df),
    ]:
        try:
            data.to_excel(paths[key], sheet_name=sheet, index=False, engine="openpyxl")
            adjust_format_excel(paths[key])
            log(f"  Saved: {os.path.basename(paths[key])}")
        except Exception as e:
            log(f"  Error saving {key} file: {e}")

    try:
        combined = pd.concat([valid_df, catch_all_df], ignore_index=True)
        combined.to_excel(paths["valid_catch_all"], sheet_name="Valid_and_Catch-All", index=False, engine="openpyxl")
        adjust_format_excel(paths["valid_catch_all"])
        log(f"  Saved: {os.path.basename(paths['valid_catch_all'])}")
    except Exception as e:
        log(f"  Error saving combined valid/catch-all file: {e}")

    log("Processing complete!")
    return paths
