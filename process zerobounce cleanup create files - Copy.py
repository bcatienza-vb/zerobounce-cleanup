import pandas as pd
import sys
import os
import numbers # Import the numbers module for numeric type checking

# --- Dependency Check ---
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side 
except ImportError:
    print("Dependency 'openpyxl' not found.")
    print("Please install it by running: pip install openpyxl")
    input("Press Enter to exit...")
    sys.exit(1)

def adjust_format_excel(excel_path, width=20, number_format_threshold=7):
    """
    Adjusts column widths, removes header borders, and converts/formats
    large numbers stored as text into actual numbers in all sheets of an Excel file.
    """
    if not os.path.exists(excel_path):
        print(f"Warning: Skipping formatting for non-existent file: {os.path.basename(excel_path)}")
        return

    print(f"Applying formatting (width, header, numbers) to: {os.path.basename(excel_path)}")
    try:
        workbook = openpyxl.load_workbook(excel_path)
        no_border = Border(left=None, right=None, top=None, bottom=None)
        integer_format = '0' 

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            if worksheet.max_row == 0 or worksheet.max_column == 0:
                continue

            # 1. Remove borders from the header row (Row 1)
            if worksheet.max_row >= 1:
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.border = no_border

            # 2. Adjust column widths and format/convert numbers in data rows
            for col_idx in range(1, worksheet.max_column + 1):
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].width = width 

                if worksheet.max_row >= 2:
                    for row_idx in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        current_value = cell.value

                        convert_to_number = False
                        numeric_value = None

                        # Scenario 1: Cell contains TEXT that looks like a long number
                        if isinstance(current_value, str) and current_value.isdigit():
                            if len(current_value) > number_format_threshold:
                                try:
                                    numeric_value = int(current_value)
                                    convert_to_number = True
                                except (ValueError, TypeError):
                                    pass

                        # Scenario 2: Cell already contains a NUMBER that is large
                        elif isinstance(current_value, numbers.Number) and not isinstance(current_value, bool):
                            try:
                                if isinstance(current_value, float):
                                    num_int_part = int(current_value)
                                else:
                                    num_int_part = current_value

                                if len(str(abs(num_int_part))) > number_format_threshold:
                                    numeric_value = current_value 
                                    convert_to_number = True 
                            except (ValueError, TypeError):
                                pass

                        if convert_to_number and numeric_value is not None:
                            cell.value = numeric_value
                            cell.number_format = integer_format

        workbook.save(excel_path)
        print(f"Successfully applied formatting to: {os.path.basename(excel_path)}")
    except Exception as e:
        print(f"Warning: Could not apply formatting to {os.path.basename(excel_path)}. Error: {e}")
        if isinstance(e, PermissionError):
            print("   Please ensure the file is not already open.")


def process_csv_to_excel(file_path):
    """
    Processes a CSV file to organize data based on 'ZB Status' and 'ZB Sub status',
    removes those columns from the final output, and saves results into 'cleaned' folder.
    """
    print(f"Processing file: {file_path}")

    # --- 1. Validate Input ---
    if not os.path.isfile(file_path):
        print(f"Error: File not found at {file_path}")
        return
    if not file_path.lower().endswith('.csv'):
        print(f"Error: File is not a CSV: {file_path}")
        return

    # --- 2. Read CSV ---
    try:
        try:
            potential_long_num_cols = ['Company Telephone Number'] 
            dtype_dict = {col: str for col in potential_long_num_cols if col in pd.read_csv(file_path, nrows=1).columns}
            df = pd.read_csv(file_path, encoding='utf-8', low_memory=False, dtype=dtype_dict)
        except UnicodeDecodeError:
            print("UTF-8 decoding failed, trying latin1 encoding...")
            dtype_dict = {col: str for col in potential_long_num_cols if col in pd.read_csv(file_path, nrows=1, encoding='latin1').columns}
            df = pd.read_csv(file_path, encoding='latin1', low_memory=False, dtype=dtype_dict)
        print("CSV file read successfully.")
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        return

    # --- 3. Find Required Columns ---
    required_cols = ["ZB Status", "ZB Sub status"]
    if not all(col in df.columns for col in required_cols):
        missing = [col for col in required_cols if col not in df.columns]
        print(f"Error: Could not find required columns: {', '.join(missing)}. Please check your data.")
        return
    print("Required columns found.")

    # --- 4. Prepare Data (Conversion and Cleaning) ---
    if "ZB Status" in df.columns:
        df['ZB Status'] = df['ZB Status'].fillna('').astype(str).str.lower().str.strip()
    if "ZB Sub status" in df.columns:
        df['ZB Sub status'] = df['ZB Sub status'].fillna('').astype(str).str.lower().str.strip()

    # --- 5. Filter Data ---
    status_col = df.get('ZB Status', pd.Series([False] * len(df), dtype=bool))
    sub_status_col = df.get('ZB Sub status', pd.Series([False] * len(df), dtype=bool))

    cond1 = (status_col == 'do_not_mail') & \
            (sub_status_col.isin(['role_based_catch_all', 'role_based']))
    cond2 = (status_col == 'valid')
    cond3 = (status_col == 'catch-all') | \
            (sub_status_col == 'catch_all')

    is_catch_all = cond1 | cond3
    is_valid = cond2 & ~is_catch_all 
    is_invalid = ~(is_valid | is_catch_all)

    # --- 6. Create Output DataFrames ---
    valid_df = df[is_valid].copy()
    catch_all_df = df[is_catch_all].copy()
    invalid_df = df[is_invalid].copy()

    print(f"\nFiltering complete: "
          f"{len(valid_df)} Valid, "
          f"{len(catch_all_df)} Catch-All, "
          f"{len(invalid_df)} Invalid rows.")

    # --- 7. Calculate Statistics ---
    stats_horizontal_data = {
        'Valid': [len(valid_df)],
        'Catch-All': [len(catch_all_df)],
        'Invalid': [len(invalid_df)],
        'Total': [len(df)]
    }
    stats_horizontal_df = pd.DataFrame(stats_horizontal_data)

    # --- 8. REMOVE ZB COLUMNS FOR FINAL OUTPUT ---
    # We do this AFTER filtering and stats, but BEFORE saving.
    cols_to_remove = ["ZB Status", "ZB Sub status"]
    
    print(f"Removing {cols_to_remove} from final output files...")
    valid_df.drop(columns=cols_to_remove, errors='ignore', inplace=True)
    catch_all_df.drop(columns=cols_to_remove, errors='ignore', inplace=True)
    invalid_df.drop(columns=cols_to_remove, errors='ignore', inplace=True)

    # --- 9. Determine and Create Output Directory/Paths ---
    input_dir = os.path.dirname(file_path)
    base_name = os.path.basename(file_path)
    file_name_without_ext, _ = os.path.splitext(base_name)

    output_base_name = file_name_without_ext
    for pattern in ["all_results", "ALL_RESULTS"]: 
         output_base_name = output_base_name.replace(pattern, "")
    output_base_name = output_base_name.strip(' _-') 
    if not output_base_name: 
        output_base_name = "Processed_Data"

    cleaned_dir = os.path.join(input_dir, "cleaned")
    try:
        os.makedirs(cleaned_dir, exist_ok=True) 
    except OSError as e:
        print(f"Error: Could not create output directory '{cleaned_dir}'. Error: {e}")
        return 

    processed_excel_path = os.path.join(cleaned_dir, f"{output_base_name}_all_results.xlsx")
    valid_excel_path = os.path.join(cleaned_dir, f"{output_base_name}_Valid.xlsx")
    catch_all_excel_path = os.path.join(cleaned_dir, f"{output_base_name}_Catch-All.xlsx")
    combined_valid_catchall_path = os.path.join(cleaned_dir, f"{output_base_name}_Valid_and_Catch-All.xlsx")

    # --- 10. Save Combined Output Excel File ---
    print(f"\nAttempting to save combined file: {processed_excel_path}")
    try:
        with pd.ExcelWriter(processed_excel_path, engine='openpyxl') as writer:
            valid_df.to_excel(writer, sheet_name='Valid', index=False)
            catch_all_df.to_excel(writer, sheet_name='Catch-All', index=False)
            invalid_df.to_excel(writer, sheet_name='Invalid', index=False)
            stats_horizontal_df.to_excel(writer, sheet_name='Statistics', index=False)
        print(f"Successfully saved combined data structure to: {processed_excel_path}")
        adjust_format_excel(processed_excel_path, width=20) 

    except Exception as e:
        print(f"\nError writing combined Excel file ({processed_excel_path}): {e}")
        if isinstance(e, PermissionError):
            print("   Please ensure the file is not already open.")

    # --- 11. Save Separate Valid Excel File ---
    print(f"\nAttempting to save separate Valid file: {valid_excel_path}")
    try:
        valid_df.to_excel(valid_excel_path, sheet_name='Valid', index=False, engine='openpyxl')
        print(f"Successfully saved Valid data structure to: {valid_excel_path}")
        adjust_format_excel(valid_excel_path, width=20) 
    except Exception as e:
        print(f"\nError writing Valid Excel file ({valid_excel_path}): {e}")

    # --- 12. Save Separate Catch-All Excel File ---
    print(f"\nAttempting to save separate Catch-All file: {catch_all_excel_path}")
    try:
        catch_all_df.to_excel(catch_all_excel_path, sheet_name='Catch-All', index=False, engine='openpyxl')
        print(f"Successfully saved Catch-All data structure to: {catch_all_excel_path}")
        adjust_format_excel(catch_all_excel_path, width=20) 
    except Exception as e:
        print(f"\nError writing Catch-All Excel file ({catch_all_excel_path}): {e}")

    # --- 13. Save Combined Valid and Catch-All Excel File (Same Sheet) ---
    print(f"\nAttempting to save combined Valid and Catch-All file: {combined_valid_catchall_path}")
    try:
        combined_valid_catchall_df = pd.concat([valid_df, catch_all_df], ignore_index=True)
        combined_valid_catchall_df.to_excel(combined_valid_catchall_path, sheet_name='Valid_and_Catch-All', index=False, engine='openpyxl')
        print(f"Successfully saved combined Valid and Catch-All data to: {combined_valid_catchall_path}")
        adjust_format_excel(combined_valid_catchall_path, width=20) 
    except Exception as e:
        print(f"\nError writing combined Valid and Catch-All Excel file ({combined_valid_catchall_path}): {e}")

    print("\nData processing complete!")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        file_path_arg = ' '.join(sys.argv[1:])
        process_csv_to_excel(file_path_arg)
    else:
        print("Usage: Drag and drop a CSV file onto this script.")
        print("Ensure you have 'pandas' and 'openpyxl' installed.")

    input("\nPress Enter to exit...")